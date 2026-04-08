import os
import argparse
import json
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
import xlrd

import datetime

def parse_time_interval(interval_str):
    """
    解析时间间隔字符串
    :param interval_str: 时间间隔字符串，如 "1y"、"12mo"、"30d"、"24h"、"60m"
    :return: 时间间隔的天数
    """
    if not interval_str:
        return None
    
    # 处理不同的时间单位
    if interval_str.endswith('mo'):
        # 月份
        try:
            value = int(interval_str[:-2])
            return value * 30
        except ValueError:
            return None
    else:
        # 其他单位
        unit = interval_str[-1].lower()
        try:
            value = int(interval_str[:-1])
        except ValueError:
            return None
        
        if unit == 'y':
            return value * 365
        elif unit == 'd':
            return value
        elif unit == 'h':
            return value / 24
        elif unit == 'm':
            return value / (24 * 60)
        else:
            return None

def is_file_within_time_range(file_path, days):
    """
    检查文件是否在指定的时间范围内
    :param file_path: 文件路径
    :param days: 时间范围（天数）
    :return: 是否在时间范围内
    """
    if days is None:
        return True
    
    try:
        file_mtime = os.path.getmtime(file_path)
        file_time = datetime.datetime.fromtimestamp(file_mtime)
        current_time = datetime.datetime.now()
        time_diff = (current_time - file_time).total_seconds() / (24 * 3600)  # 转换为天数
        return time_diff <= days
    except Exception:
        return False

def find_excel_files(directory, filename_pattern=None, max_files=None, time_interval=None):
    """
    查找目录及其子目录下的所有Excel文件
    :param directory: 要搜索的目录
    :param filename_pattern: 文件名关键词筛选
    :param max_files: 最大文件数量限制
    :param time_interval: 时间间隔，如 "1y"、"12m"
    :return: Excel文件路径列表
    """
    excel_files = []
    time_days = parse_time_interval(time_interval)
    
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(('.xlsx', '.xls')):
                file_path = os.path.join(root, file)
                
                # 检查时间范围
                if not is_file_within_time_range(file_path, time_days):
                    continue
                
                # 检查文件名筛选
                if filename_pattern:
                    if fuzz.partial_ratio(filename_pattern.lower(), file.lower()) > 60:
                        excel_files.append(file_path)
                else:
                    excel_files.append(file_path)
            
            # 检查是否达到文件数量限制
            if max_files and len(excel_files) >= max_files:
                return excel_files
    return excel_files

def search_excel_content(file_path, keyword, fuzzy=False, threshold=80):
    """
    搜索Excel文件内容
    :param file_path: Excel文件路径
    :param keyword: 搜索关键词
    :param fuzzy: 是否启用模糊搜索
    :param threshold: 模糊搜索匹配阈值
    :return: 搜索结果列表
    """
    results = []
    try:
        if file_path.endswith('.xlsx'):
            # 处理xlsx文件
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell is not None:
                            cell_str = str(cell)
                            if fuzzy:
                                # 使用模糊搜索
                                if fuzz.partial_ratio(keyword.lower(), cell_str.lower()) > threshold:
                                    results.append({
                                        'file': file_path,
                                        'sheet': sheet_name,
                                        'row': row[0] if row else '',
                                        'row_num': row_idx,
                                        'col_num': col_idx,
                                        'content': cell_str
                                    })
                            else:
                                # 使用精确匹配
                                if keyword.lower() in cell_str.lower():
                                    results.append({
                                        'file': file_path,
                                        'sheet': sheet_name,
                                        'row': row[0] if row else '',
                                        'row_num': row_idx,
                                        'col_num': col_idx,
                                        'content': cell_str
                                    })
        elif file_path.endswith('.xls'):
            # 处理xls文件
            wb = xlrd.open_workbook(file_path)
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    for col_idx, cell in enumerate(row, start=1):
                        if cell is not None:
                            cell_str = str(cell)
                            if fuzzy:
                                # 使用模糊搜索
                                if fuzz.partial_ratio(keyword.lower(), cell_str.lower()) > threshold:
                                    results.append({
                                        'file': file_path,
                                        'sheet': sheet_name,
                                        'row': row[0] if row else '',
                                        'row_num': row_idx + 1,
                                        'col_num': col_idx,
                                        'content': cell_str
                                    })
                            else:
                                # 使用精确匹配
                                if keyword.lower() in cell_str.lower():
                                    results.append({
                                        'file': file_path,
                                        'sheet': sheet_name,
                                        'row': row[0] if row else '',
                                        'row_num': row_idx + 1,
                                        'col_num': col_idx,
                                        'content': cell_str
                                    })
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {e}")
    return results

def save_results(results, output_file):
    """
    保存搜索结果到本地文件
    :param results: 搜索结果列表
    :param output_file: 输出文件路径
    """
    # 转换文件路径为单反斜杠格式
    for result in results:
        if 'file' in result:
            result['file'] = result['file'].replace('\\', '/')
    
    # 按文件分组结果
    grouped_results = {}
    for result in results:
        file_path = result['file']
        if file_path not in grouped_results:
            grouped_results[file_path] = []
        grouped_results[file_path].append(result)
    
    # 构建分组后的结果结构
    structured_results = {
        'total_matches': len(results),
        'total_files': len(grouped_results),
        'files': []
    }
    
    for file_path, file_results in grouped_results.items():
        # 按工作表分组
        sheet_results = {}
        for result in file_results:
            sheet_name = result['sheet']
            if sheet_name not in sheet_results:
                sheet_results[sheet_name] = []
            sheet_results[sheet_name].append(result)
        
        file_info = {
            'file_path': file_path,
            'match_count': len(file_results),
            'sheets': []
        }
        
        for sheet_name, sheet_items in sheet_results.items():
            sheet_info = {
                'sheet_name': sheet_name,
                'match_count': len(sheet_items),
                'matches': sheet_items
            }
            file_info['sheets'].append(sheet_info)
        
        structured_results['files'].append(file_info)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(structured_results, f, ensure_ascii=False, indent=2)
    print(f"\n搜索结果已保存到: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Excel文件内容搜索工具')
    parser.add_argument('directory', help='要搜索的目录')
    parser.add_argument('keyword', help='搜索关键词')
    parser.add_argument('--filename', help='文件名筛选关键词')
    parser.add_argument('--max-files', type=int, help='最大搜索文件数量')
    parser.add_argument('--output', help='保存搜索结果到文件')
    parser.add_argument('--fuzzy', action='store_true', help='启用模糊搜索')
    parser.add_argument('--threshold', type=int, default=80, help='模糊搜索匹配阈值 (默认: 80)')
    parser.add_argument('--time', help='时间间隔，如 "1y"、"12m"、"30d"')
    
    args = parser.parse_args()
    
    print(f"开始搜索目录: {args.directory}")
    print(f"搜索关键词: {args.keyword}")
    if args.filename:
        print(f"文件名筛选: {args.filename}")
    if args.max_files:
        print(f"最大文件数量: {args.max_files}")
    if args.time:
        print(f"时间间隔: {args.time}")
    if args.fuzzy:
        print(f"模糊搜索: 启用 (阈值: {args.threshold})")
    else:
        print(f"模糊搜索: 禁用 (精确匹配)")
    
    # 查找Excel文件
    excel_files = find_excel_files(args.directory, args.filename, args.max_files, args.time)
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    # 搜索内容
    all_results = []
    for file_path in excel_files:
        print(f"正在搜索文件: {file_path}")
        results = search_excel_content(file_path, args.keyword, args.fuzzy, args.threshold)
        all_results.extend(results)
    
    # 按文件分组结果
    grouped_results = {}
    for result in all_results:
        file_path = result['file']
        if file_path not in grouped_results:
            grouped_results[file_path] = []
        grouped_results[file_path].append(result)
    
    # 输出结果
    print(f"\n搜索完成，共找到 {len(all_results)} 个匹配项")
    print(f"分布在 {len(grouped_results)} 个文件中")
    
    file_index = 1
    for file_path, file_results in grouped_results.items():
        print(f"\n{file_index}. 文件: {file_path}")
        print(f"   匹配项数量: {len(file_results)}")
        
        # 按工作表分组
        sheet_results = {}
        for result in file_results:
            sheet_name = result['sheet']
            if sheet_name not in sheet_results:
                sheet_results[sheet_name] = []
            sheet_results[sheet_name].append(result)
        
        for sheet_name, sheet_items in sheet_results.items():
            print(f"   工作表: {sheet_name} ({len(sheet_items)}个匹配项)")
            for item_idx, item in enumerate(sheet_items, 1):
                print(f"      {item_idx}. 位置: 第{item['row_num']}行, 第{item['col_num']}列")
                print(f"         行标识: {item['row']}")
                print(f"         内容: {item['content']}")
        
        file_index += 1
    
    # 保存结果
    if args.output:
        save_results(all_results, args.output)

if __name__ == '__main__':
    main()
    # python main.py tongji D:\all_projects --output test_results.json --time 1y
    # python main.py D:\all_projects tongji --filename Profile --max-files 1000 --output search_results.json --fuzzy --threshold 95 --time 1y